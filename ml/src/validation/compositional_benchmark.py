"""
Validate the tier-2 compositional calculator against INDB's own published
dish values.

WHY THIS IS A STRONG TEST:
INDB publishes BOTH the finished nutrition of 1,014 Indian dishes AND the
ingredient list each was computed from. So the calculator can be handed
the ingredients and asked to reproduce the published dish. If it does, the
whole chain works: unit conversion, density, ingredient lookup in our own
database, and cooking yield.

Crucially this is NOT circular. The calculator resolves each ingredient
through OUR search against OUR unified database (IFCT/USDA/CNF/OFF) -- it
never reads INDB's per-ingredient values. So agreement means our
ingredient data and unit handling independently reproduce a published
recipe calculation, which is exactly the capability tier 2 needs for a
dish nobody has published at all.

WHAT DISAGREEMENT WOULD MEAN:
  * unit conversion wrong        -> large, systematic, direction-consistent
  * ingredient lookup wrong      -> large errors on specific dishes
  * yield factors wrong          -> per-100g wrong, TOTALS still right
So per-serving and per-100g are reported separately; that split localises
the fault instead of yielding one uninterpretable number.
"""
import json
import sys
from pathlib import Path
from collections import defaultdict

import numpy as np
import openpyxl

SRC = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(SRC))

from inference.compositional import CompositionalCalculator  # noqa: E402

RAW = Path(__file__).resolve().parents[2] / "data" / "raw" / "food_v1" / "indb" / \
    "Indian-Nutrient-Databank-INDB--main"
PROC = Path(__file__).resolve().parents[2] / "data" / "processed"
OUT = PROC / "compositional_benchmark.json"


def load_recipes():
    wb = openpyxl.load_workbook(RAW / "recipes.xlsx", read_only=True)
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows))
    col = {n: i for i, n in enumerate(hdr)}
    by_code = defaultdict(list)
    names = {}
    for r in rows:
        code = r[col["recipe_code"]]
        if not code:
            continue
        names[code] = r[col["recipe_name"]]
        by_code[code].append({
            # Use the recipe's own free-text ingredient name, NOT the mapped
            # food_name -- the mapped column is INDB's answer, and feeding it
            # back would test nothing about our own resolution.
            "name": str(r[col["ingredient_name_org"]] or "").strip(),
            "amount": r[col["amount_org"]],
            "unit": r[col["unit_org"]],
        })
    return by_code, names


def load_servings():
    wb = openpyxl.load_workbook(RAW / "recipes_servingsize.xlsx", read_only=True)
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows))
    col = {n: i for i, n in enumerate(hdr)}
    out = {}
    for r in rows:
        code = r[col["recipe_code"]]
        if not code:
            continue
        try:
            n = float(r[col["no_of_servings"]])
            if n > 0:
                out[code] = n
        except (TypeError, ValueError):
            pass
    return out


def ape(pred, truth):
    if pred is None or truth is None or truth <= 0:
        return None
    return abs(pred - truth) / truth * 100.0


def main():
    recipes, names = load_recipes()
    servings = load_servings()
    published = {d["source_id"].split(":", 1)[1]: d
                 for d in json.loads((PROC / "indb_dishes.json").read_text(encoding="utf-8"))}

    calc = CompositionalCalculator()
    rows = []
    print(f"recipes to evaluate: {len(recipes)}", flush=True)

    for i, (code, ings) in enumerate(recipes.items(), 1):
        if i % 200 == 0:
            print(f"  {i}/{len(recipes)}", flush=True)
        truth = published.get(code)
        if not truth or truth.get("data_quality_flag"):
            continue
        n_serv = servings.get(code)
        if not n_serv:
            continue
        res = calc.compute(ings, servings=n_serv, dish_name=names.get(code))
        if not res.get("ok"):
            rows.append({"code": code, "name": names.get(code), "ok": False})
            continue

        t_serv_kcal = truth.get("serving_energy_kcal")
        p_serv_kcal = (res.get("per_serving") or {}).get("energy_kcal")
        t_100 = truth.get("energy_kcal")
        p_100 = (res.get("per_100g_cooked") or {}).get("energy_kcal")

        rows.append({
            "code": code, "name": names.get(code), "ok": True,
            "confidence": res.get("confidence"),
            "ingredients_used": res["ingredients_used"],
            "unresolved": len(res["unresolved"]),
            "pred_serving_kcal": p_serv_kcal, "truth_serving_kcal": t_serv_kcal,
            "ape_serving": ape(p_serv_kcal, t_serv_kcal),
            "pred_100g_kcal": p_100, "truth_100g_kcal": t_100,
            "ape_100g": ape(p_100, t_100),
            "pred_protein_serv": (res.get("per_serving") or {}).get("protein_g"),
        })

    ok = [r for r in rows if r.get("ok")]
    OUT.write_text(json.dumps(rows, indent=2), encoding="utf-8")

    def summarise(key, label):
        v = [r[key] for r in ok if r.get(key) is not None]
        if not v:
            print(f"  {label}: no comparable rows")
            return None
        a = np.array(v)
        print(f"  {label:26s} n={len(a):4d}  median {np.median(a):5.1f}%  "
              f"within25% {(a <= 25).mean()*100:5.1f}%  within50% {(a <= 50).mean()*100:5.1f}%")
        return {"n": len(a), "median_ape": round(float(np.median(a)), 1),
                "within_25": round(float((a <= 25).mean() * 100), 1),
                "within_50": round(float((a <= 50).mean() * 100), 1)}

    print(f"\nTIER 2 vs INDB PUBLISHED VALUES  ({len(ok)} dishes computed)")
    s_serv = summarise("ape_serving", "per-serving energy")
    s_100 = summarise("ape_100g", "per-100g energy")

    print("\n  by calculator confidence (per-serving energy):")
    for conf in ("high", "medium", "low"):
        v = [r["ape_serving"] for r in ok
             if r.get("confidence") == conf and r.get("ape_serving") is not None]
        if v:
            a = np.array(v)
            print(f"    {conf:8s} n={len(a):4d}  median {np.median(a):5.1f}%  "
                  f"within25% {(a <= 25).mean()*100:5.1f}%")

    worst = sorted((r for r in ok if r.get("ape_serving") is not None),
                   key=lambda r: -r["ape_serving"])[:10]
    print("\n  worst per-serving mismatches (diagnostic):")
    for r in worst:
        print(f"    {r['ape_serving']:7.0f}%  {str(r['name'])[:40]:40s} "
              f"pred {r['pred_serving_kcal']:7.1f} vs {r['truth_serving_kcal']:7.1f}")

    summary = {"dishes_computed": len(ok), "per_serving": s_serv, "per_100g": s_100}
    (PROC / "compositional_benchmark_summary.json").write_text(
        json.dumps(summary, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
