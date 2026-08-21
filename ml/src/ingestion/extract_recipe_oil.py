"""
Extract the BASELINE cooking-fat quantity already baked into every INDB
dish, so a user's oil selection can adjust from a real starting point
rather than from a guess.

WHY THIS IS THE RIGHT LEVER:
The measured irreducible error for name-only prediction is ~17% median,
and the single largest driver is cooking fat -- a samosa swings ~2x on
oil absorption, a curry ~3x on ghee and cream. That variance is not
recoverable from the dish NAME, but it IS recoverable if the user simply
tells us how much oil they used. This converts the largest unmeasurable
term into a measured input.

CRITICAL PREREQUISITE (the reason this file exists):
An oil adjustment is meaningless without knowing what oil is ALREADY in
the published number. INDB dishes are computed from recipes that already
include a specific oil quantity. If a user says "high oil" and we simply
add calories, we double-count whatever the recipe already assumed. So the
recipe's own oil content must be extracted per dish and used as the
baseline that user selections adjust FROM.

UNITS: recipe amounts are in g / tsp / tbsp / cup / ml. Converted with
standard culinary volumes and the density of cooking oil (0.92 g/ml).
Conversions are stated explicitly below rather than hidden in a dict, so
they can be checked.
"""
import json
import re
from pathlib import Path
from collections import defaultdict

import openpyxl

RAW = Path(__file__).resolve().parents[2] / "data" / "raw" / "food_v1" / "indb" / \
    "Indian-Nutrient-Databank-INDB--main"
PROC = Path(__file__).resolve().parents[2] / "data" / "processed"
OUT_PATH = PROC / "indb_recipe_oil.json"

# Density of common cooking oils/ghee ~0.92 g/ml; butter ~0.91 g/ml.
# One figure is used for all cooking fats -- the difference is <2%, far
# below the variance this feature exists to address.
OIL_DENSITY_G_PER_ML = 0.92

# Standard culinary volumes in ml. Indian recipe convention.
ML_PER = {
    "tsp": 5.0, "teaspoon": 5.0,
    "tbsp": 15.0, "tablespoon": 15.0,
    "cup": 240.0, "cups": 240.0,
    "ml": 1.0,
}
# Fats measured in vague units. A "knob" of butter is conventionally
# ~15g; anything else vague is skipped rather than guessed.
FIXED_GRAMS = {"knob": 15.0, "blob": 10.0}

# Cooking fat is detected on the MAPPED food name (the `food_name` column),
# not the recipe's free-text ingredient name. INDB frequently writes the
# ingredient as literally "Fat" -- searching the free-text name alone missed
# every one of those and reported paratha, puri and countless fried dishes
# as containing zero oil. The mapped column resolves "Fat" -> "Oil,
# sunflower", which is authoritative.
FAT_RE = re.compile(
    r"\b(oil|ghee|butter|vanaspati|margarine|dalda|shortening|fat)\b", re.I)

# Unquantified frying instructions -- "for frying", "as required", "enough".
# These carry no number, so the oil cannot be measured from the recipe.
# Treated as a DEEP-FRY MARKER rather than as zero: recording zero would
# label a deep-fried food as oil-free, which is the worst possible error
# for this feature.
UNQUANTIFIED_RE = re.compile(
    r"(for\s+fry|frying|deep\s*fry|as\s+required|as\s+needed|enough|to\s+taste|"
    r"as\s+per|greasing|shallow\s*fry)", re.I)
# Exclude ingredients that merely CONTAIN the word but are not cooking fat
# added by the cook (e.g. "buttermilk" is not butter, "butter beans" is a
# legume). Without this the baseline is inflated and every adjustment is
# wrong in the same direction.
NOT_FAT_RE = re.compile(
    r"\b(buttermilk|butter\s*milk|butter\s*bean|butter\s*fruit|"
    r"peanut\s*butter|nut\s*butter|almond\s*butter|cocoa\s*butter|"
    r"butter\s*paper|body\s*butter)\b", re.I)


def to_grams(amount, unit):
    """Convert a recipe amount to grams of fat. Returns None when the
    unit is unusable -- never a fabricated fallback."""
    if amount is None:
        return None
    try:
        amt = float(str(amount).strip())
    except (TypeError, ValueError):
        return None
    if amt <= 0:
        return None

    u = re.sub(r"\(.*?\)", "", str(unit or "")).strip().lower()
    u = re.sub(r"[^a-z]", "", u)

    if u in ("g", "gram", "grams", "gm", "gms"):
        return amt
    if u in FIXED_GRAMS:
        return amt * FIXED_GRAMS[u]
    if u in ML_PER:
        return amt * ML_PER[u] * OIL_DENSITY_G_PER_ML
    if u == "":
        # blank unit is ambiguous -- skip rather than assume grams
        return None
    return None


def main():
    wb = openpyxl.load_workbook(RAW / "recipes.xlsx", read_only=True)
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows))
    col = {n: i for i, n in enumerate(hdr)}

    per_recipe = defaultdict(lambda: {"oil_g": 0.0, "items": [], "unparsed": [], "deep_fry_marker": False})
    recipe_names = {}

    for r in rows:
        code = r[col["recipe_code"]]
        if not code:
            continue
        recipe_names[code] = r[col["recipe_name"]]
        ing = str(r[col["ingredient_name_org"]] or "")
        mapped = str(r[col["food_name"]] or "")

        # Decide on the MAPPED name first (authoritative), falling back to
        # the free-text name. Exclusions are checked against both, so
        # "peanut butter" and "buttermilk" never count as cooking fat.
        is_fat = bool(FAT_RE.search(mapped) or FAT_RE.search(ing))
        if not is_fat or NOT_FAT_RE.search(mapped) or NOT_FAT_RE.search(ing):
            continue

        amount_txt = f"{r[col['amount_org']]} {r[col['unit_org']]}"
        grams = to_grams(r[col["amount_org"]], r[col["unit_org"]])
        entry = {"ingredient": ing.strip(), "mapped": mapped.strip(),
                 "amount": r[col["amount_org"]], "unit": r[col["unit_org"]]}

        if grams is None:
            if UNQUANTIFIED_RE.search(amount_txt):
                # deep-fry / greasing marker: real oil, quantity unknowable
                per_recipe[code]["deep_fry_marker"] = True
            per_recipe[code]["unparsed"].append(entry)
        else:
            per_recipe[code]["oil_g"] += grams
            per_recipe[code]["items"].append({**entry, "grams": round(grams, 1)})

    # Servings per recipe -- REQUIRED, because recipe oil is the amount for
    # the WHOLE recipe. Dividing by servings is the only way to get the oil
    # actually present in one portion; skipping this step would overstate
    # oil by the serving count (often 4-6x).
    sw = openpyxl.load_workbook(RAW / "recipes_servingsize.xlsx", read_only=True)["Sheet1"]
    srows = sw.iter_rows(values_only=True)
    shdr = list(next(srows))
    scol = {n: i for i, n in enumerate(shdr)}
    servings_by_code = {}
    for r in srows:
        code = r[scol["recipe_code"]]
        if not code:
            continue
        try:
            n = float(r[scol["no_of_servings"]])
            if n > 0:
                servings_by_code[code] = n
        except (TypeError, ValueError):
            pass

    # join to per-dish nutrition to express oil per 100g and per serving
    dishes = {d["source_id"].split(":", 1)[1]: d
              for d in json.loads((PROC / "indb_dishes.json").read_text(encoding="utf-8"))}

    out = []
    for code, info in per_recipe.items():
        d = dishes.get(code)
        if not d:
            continue
        total_oil = round(info["oil_g"], 1)
        serving_g = d.get("serving_grams")
        n_serv = servings_by_code.get(code)

        oil_per_serving = round(total_oil / n_serv, 2) if n_serv else None
        oil_per_100g = None
        if oil_per_serving is not None and serving_g:
            oil_per_100g = round(oil_per_serving / serving_g * 100, 2)

        rec = {
            "food_code": code,
            "food_name": recipe_names.get(code) or d.get("food_name"),
            "recipe_total_oil_g": total_oil,
            "no_of_servings": n_serv,
            "oil_g_per_serving": oil_per_serving,
            "oil_g_per_100g": oil_per_100g,
            "oil_ingredients": info["items"],
            "unparsed_oil_ingredients": info["unparsed"],
            "deep_fried_unquantified": info.get("deep_fry_marker", False),
            "energy_kcal_per_100g": d.get("energy_kcal"),
            "serving_grams": serving_g,
        }

        # DEEP-FRYING CAVEAT: a recipe listing 300-500 g of oil is specifying
        # a frying BATH, not oil that ends up in the food -- typical
        # absorption is 5-15% of the food's weight, not the whole bath.
        # Treating bath oil as consumed oil would overstate a plate of
        # pakoras by hundreds of calories, so these are flagged and their
        # per-serving figure is marked unreliable rather than used blindly.
        if oil_per_100g is not None and oil_per_100g > 25:
            rec["oil_baseline_unreliable"] = (
                f"implied {oil_per_100g} g oil per 100 g of finished dish, which "
                "exceeds plausible absorption -- the recipe almost certainly "
                "specifies a deep-frying bath rather than oil retained in the food"
            )
        out.append(rec)

    # Recipes with NO fat ingredient at all are informative too -- they are
    # the genuine zero-oil baseline (boiled, steamed, raw dishes).
    zero = [c for c in dishes if c not in per_recipe]
    OUT_PATH.write_text(json.dumps(
        {"with_oil": out, "no_oil_recipe_codes": zero}, indent=2), encoding="utf-8")

    parsed = [r for r in out if r["recipe_total_oil_g"] > 0]
    unparsed_count = sum(len(r["unparsed_oil_ingredients"]) for r in out)
    print(f"Dishes with a parsed cooking-fat quantity: {len(parsed)}")
    print(f"Dishes with fat listed but unusable units: {unparsed_count} ingredient rows")
    print(f"Dishes with no cooking fat in the recipe:  {len(zero)}")

    vals = sorted(r["recipe_total_oil_g"] for r in parsed)
    if vals:
        import statistics
        print(f"\nRecipe total oil (g per WHOLE recipe, all servings):")
        print(f"  median {statistics.median(vals):.1f}   p25 {vals[len(vals)//4]:.1f}   "
              f"p75 {vals[3*len(vals)//4]:.1f}   max {vals[-1]:.1f}")

    usable = [r for r in parsed
              if r.get("oil_g_per_100g") is not None and not r.get("oil_baseline_unreliable")]
    flagged = [r for r in parsed if r.get("oil_baseline_unreliable")]
    print(f"\nDishes with a usable oil-per-100g baseline: {len(usable)}")
    print(f"Flagged as deep-fry bath (baseline not usable): {len(flagged)}")

    if usable:
        import statistics
        v = sorted(r["oil_g_per_100g"] for r in usable)
        print(f"\nOIL PER 100 g OF FINISHED DISH -- the real distribution:")
        for q, lbl in ((0.10, "p10"), (0.25, "p25"), (0.50, "median"),
                       (0.75, "p75"), (0.90, "p90")):
            i = min(int(q * len(v)), len(v) - 1)
            print(f"  {lbl:>6s}  {v[i]:5.2f} g/100g   ({v[i]*9:5.1f} kcal per 100 g)")
        print(f"  {'max':>6s}  {v[-1]:5.2f} g/100g")
        print(f"\n  mean {statistics.mean(v):.2f}   n={len(v)}")

    print(f"\n  examples:")
    for r in usable[:10]:
        print(f"    {r['food_name'][:42]:42s} {r['oil_g_per_100g']:5.2f} g/100g  "
              f"({r['oil_g_per_serving']:5.1f} g per {r['serving_grams']:.0f} g serving)")


if __name__ == "__main__":
    main()
