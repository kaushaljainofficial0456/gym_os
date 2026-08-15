"""
Ingest Reis et al. 2017, PLOS ONE (DOI 10.1371/journal.pone.0181311) —
"Energy cost of isolated resistance exercises across low- to high-intensities".
CC BY 4.0. Source file: ml/data/external/reis2017_pone.0181311.s001.xlsx

RAW SCHEMA (discovered by direct inspection, not assumed):
Each sheet = one exercise. Each sheet has TWO stacked 14-row blocks, marked
by cell fill color, both re-indexed 1..14 as the participant number:

  * Rows 2-15  (fill FFFFFF00, "yellow"): raw relative VO2, ml.kg-1.min-1,
    one column per %1RM intensity (12/16/20/24/80). This is the sheet's own
    stated ground truth ("yellow is VO2 (ml/g.min)" — the header note's "g"
    is treated as a typo/shorthand for kg; see verification below).
  * Rows 17-30 (fill FF00B0F0, "blue"): the SAME 14 participants' data
    converted to kcal/min — but via the file's own within-sheet arithmetic
    using the COHORT MEAN body weight (78.67 kg, matching the paper's
    reported mean +/- SD), not each participant's individual weight. This
    was confirmed empirically: yellow/blue ratio is a CONSTANT 2.542 across
    every row and every exercise checked, and 1000 / (2.542 * 5) = 78.68 kg
    — matching the paper's cohort mean exactly. So the blue "kcal/min"
    column is a derived, group-level approximation baked into the file,
    NOT an individually accurate label. It is therefore NOT used as the
    ground-truth target here; the yellow relative-VO2 values are used
    instead (see harmonize.py for the individual-weight problem this
    creates, since Reis2017 has no per-participant weight of its own).

CROSS-STUDY FINDING (important for leakage prevention): the blue block's
kcal/min values are numerically identical (to several decimal places) to
Reis et al. 2019's directly-reported "EC" (energy cost) columns for the
same exercises/intensities. This is strong, direct evidence — not just
demographic similarity — that Reis 2017 and Reis 2019 share the same
underlying ~14-17 participant cohort. harmonize.py treats every row from
both files as belonging to one shared participant_group_id and NEVER lets
a train/val/test split separate them.
"""
from pathlib import Path
import openpyxl
import pandas as pd

from ontology.exercise_map import to_canonical

RAW_PATH = Path(__file__).resolve().parents[2] / "data" / "external" / "reis2017_pone.0181311.s001.xlsx"

YELLOW = "FFFFFF00"
BLUE = "FF00B0F0"

# Cohort mean body weight as reported in the paper text (Reis et al. 2017).
# Used ONLY to document/verify the blue block's derivation — never to
# back-convert a per-participant kcal/min value, since that would just
# re-inject the same group-mean approximation this file already made.
COHORT_MEAN_WEIGHT_KG = 78.67
COHORT_SD_WEIGHT_KG = 10.7


def _sheet_block(ws, start_row, header_row, n_rows=14):
    headers = [ws.cell(row=header_row, column=c).value for c in range(2, ws.max_column + 1)]
    rows = []
    for r in range(start_row, start_row + n_rows):
        pid = ws.cell(row=r, column=1).value
        if pid is None:
            continue
        for c, h in zip(range(2, ws.max_column + 1), headers):
            if h is None:
                continue
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            rows.append({"local_participant_id": int(pid), "column_label": str(h), "value": float(val)})
    return rows


def load() -> pd.DataFrame:
    wb = openpyxl.load_workbook(RAW_PATH, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        yellow_rows = _sheet_block(ws, start_row=2, header_row=1, n_rows=14)
        blue_rows = _sheet_block(ws, start_row=17, header_row=16, n_rows=14)

        canonical = to_canonical(sheet_name)

        for row in yellow_rows:
            pct = "".join(ch for ch in row["column_label"].split("-")[-1] if ch.isdigit())
            out.append({
                "dataset_id": "reis2017",
                "participant_group_id": f"reis_lab_p{row['local_participant_id']}",
                "local_participant_id": row["local_participant_id"],
                "exercise_canonical_id": canonical,
                "exercise_original_label": sheet_name,
                "condition_type": "pct_1rm",
                "intensity_value": int(pct) if pct else None,
                "metric_type": "vo2_relative",
                "metric_subtype": None,
                "value": row["value"],
                "unit": "ml_kg_min",
                "is_directly_measured": True,
                "is_group_mean_derived": False,
                "source_sheet": sheet_name,
                "notes": None,
            })
        for row in blue_rows:
            pct = "".join(ch for ch in row["column_label"].split("-")[-1] if ch.isdigit())
            out.append({
                "dataset_id": "reis2017",
                "participant_group_id": f"reis_lab_p{row['local_participant_id']}",
                "local_participant_id": row["local_participant_id"],
                "exercise_canonical_id": canonical,
                "exercise_original_label": sheet_name,
                "condition_type": "pct_1rm",
                "intensity_value": int(pct) if pct else None,
                "metric_type": "energy_cost_rate",
                "metric_subtype": None,
                "value": row["value"],
                "unit": "kcal_min",
                "is_directly_measured": False,
                "is_group_mean_derived": True,  # derived using COHORT_MEAN_WEIGHT_KG, not individual weight
                "source_sheet": sheet_name,
                "notes": (
                    f"Derived in-source from VO2 using cohort mean weight "
                    f"({COHORT_MEAN_WEIGHT_KG} kg), not this participant's own weight. "
                    f"Prefer Reis2019's directly-reported EC for the same participant/exercise/intensity "
                    f"where available (see harmonize.py dedup logic)."
                ),
            })
    return pd.DataFrame(out)


if __name__ == "__main__":
    df = load()
    print(df.shape)
    print(df.head(10))
    print(df["metric_type"].value_counts())
    print(df["exercise_canonical_id"].value_counts())
