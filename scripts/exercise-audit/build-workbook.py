#!/usr/bin/env python3
"""Assemble the exercise-library audit CSVs into one .xlsx deliverable.
No formulas — pure data. Run after inventory.mjs / dedup.mjs / build-master.mjs.
"""
import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
D = ROOT / "docs" / "exercise-library-audit"
OUT = D / "exercise-master.xlsx"

SHEETS = [
    ("Master",                  "exercise-master.csv"),
    ("Inventory (existing 207)", "inventory.csv"),
    ("Candidates + Class",       "candidates.csv"),
    ("Duplicate-Merge-Report",   "duplicate-merge-report.csv"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
NEW_FILL = PatternFill("solid", fgColor="FFF3CD")  # highlight NEW / non-KEEP rows

with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
    for sheet, csv in SHEETS:
        df = pd.read_csv(D / csv, dtype=str).fillna("")
        df.to_excel(xw, sheet_name=sheet[:31], index=False)
        ws = xw.sheets[sheet[:31]]
        # header
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        # body font + width + row highlight
        status_col = None
        for c in range(1, ws.max_column + 1):
            h = str(ws.cell(row=1, column=c).value or "").lower()
            if h in ("source", "classification", "status") and status_col is None and h != "status":
                status_col = c
            if h == "source":  # master sheet NEW/EXISTING
                status_col = c
        widths = {}
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = BODY_FONT
                v = "" if cell.value is None else str(cell.value)
                widths[c] = min(60, max(widths.get(c, 10), len(v) + 2, len(str(ws.cell(row=1, column=c).value)) + 2))
            if status_col:
                sv = str(ws.cell(row=r, column=status_col).value or "")
                if sv in ("NEW", "ALIAS_TO_EXISTING", "DUPLICATE_DO_NOT_IMPORT"):
                    for c in range(1, ws.max_column + 1):
                        ws.cell(row=r, column=c).fill = NEW_FILL
        for c, w in widths.items():
            ws.column_dimensions[get_column_letter(c)].width = w

print(f"wrote {OUT}")
for sheet, csv in SHEETS:
    n = len(pd.read_csv(D / csv))
    print(f"  {sheet:26s} {n:4d} rows  <- {csv}")
